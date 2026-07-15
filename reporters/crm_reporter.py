#!/usr/bin/env python3
"""
Репортер - генерирует HTML отчеты из данных БД
"""
import html as html_module
import io
import json
import logging
import psycopg2  # pyright: ignore[reportMissingModuleSource]
from datetime import datetime, timedelta
from flask import Flask, request, Response  # pyright: ignore[reportMissingImports]
from config import DB_CONFIG, MANAGER_NAMES, STATUS_LABELS, STATUS_CLASSES, CRM_CONFIG
from last_comment import (
    get_last_comment_date_plain,
    get_last_comment_for_display,
    get_last_comment_plain,
)

app = Flask(__name__)
logger = logging.getLogger(__name__)


def get_requests_from_db(start_date, end_date):
    """Заявки, созданные в CRM в выбранном периоде (по дате создания в БД = дата заявки из CRM)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(
        """
        SELECT data FROM requests
        WHERE created_at >= %s AND created_at <= %s
        ORDER BY created_at DESC
        """,
        (start_date, end_date),
    )

    results = cur.fetchall()
    cur.close()
    conn.close()

    # Фильтруем None; jsonb может прийти как dict или как строка JSON
    valid_data = []
    for row in results:
        raw = row[0]
        if raw is None:
            continue
        if isinstance(raw, dict):
            valid_data.append(raw)
        elif isinstance(raw, str):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, dict):
                    valid_data.append(parsed)
            except (json.JSONDecodeError, TypeError):
                continue

    return valid_data


def get_last_db_update_in_period(start_date, end_date):
    """Максимальное updated_at среди заявок, созданных в периоде (насколько свежи данные в БД)."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT MAX(updated_at) FROM requests
            WHERE created_at >= %s AND created_at <= %s
            """,
            (start_date, end_date),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row[0]:
            return row[0]
    except Exception:
        logger.debug("get_last_db_update_in_period failed", exc_info=True)
    return None


def _empty_analytics_result():
    """Та же структура, что и у analyze_data при наличии данных (нужна при пустом периоде)."""
    status_order = [
        'DRAFT',
        'NEW',
        'INCOME_CREATED',
        'INCOME_PAID',
        'INCOME_PARTIALLY_PAID',
        'KKT_LINKED',
        'COMPLETED',
        'REFUND',
        'CANCELED_BY_CLIENT',
        'CANCELED_BY_BANK',
        'ARCHIVE',
    ]
    return {
        'total_orders': 0,
        'paid_orders': 0,
        'invoiced_orders': 0,
        'conversion_rate': 0.0,
        'status_stats': {s: 0 for s in status_order},
        'manager_stats': {},
        'cancellation_rating': [],
        'total_cancellations': 0,
        'daily_stats': {},
        'raw_data': [],
    }


def analyze_data(data):
    """Анализ данных"""
    if not data:
        return _empty_analytics_result()

    # Фильтруем валидные записи
    valid_data = [order for order in data if order and isinstance(order, dict)]

    if not valid_data:
        return _empty_analytics_result()

    total_orders = len(valid_data)
    paid_and_processed_orders = len([
        order for order in valid_data
        if order.get('status') in ['INCOME_PAID', 'INCOME_PARTIALLY_PAID', 'KKT_LINKED', 'COMPLETED']
    ])
    invoiced_orders = len([order for order in valid_data
                           if order.get('status') == 'INCOME_CREATED'])
    conversion_rate = (paid_and_processed_orders / total_orders * 100) if total_orders > 0 else 0

    # Статистика по статусам
    status_order = [
        'DRAFT',
        'NEW',
        'INCOME_CREATED',
        'INCOME_PAID',
        'INCOME_PARTIALLY_PAID',
        'KKT_LINKED',
        'COMPLETED',
        'REFUND',
        'CANCELED_BY_CLIENT',
        'CANCELED_BY_BANK',
        'ARCHIVE',
    ]
    status_stats = {}
    for status in status_order:
        status_stats[status] = len([order for order in valid_data if order.get('status') == status])

    # Рейтинг причин отмен: группируем комментарии для клиентских и банковских отмен.
    cancellation_statuses = {'CANCELED_BY_CLIENT', 'CANCELED_BY_BANK'}
    cancellation_comment_stats = {}
    total_cancellations = 0
    for order in valid_data:
        status = order.get('status')
        if status not in cancellation_statuses:
            continue
        total_cancellations += 1
        comment = get_last_comment_plain(order) or 'Без комментария'
        key = (status, comment)
        cancellation_comment_stats[key] = cancellation_comment_stats.get(key, 0) + 1

    cancellation_rating = [
        {
            'status': STATUS_LABELS.get(status, status),
            'comment': comment,
            'count': count,
            'percent': round(count / total_cancellations * 100, 1),
        }
        for (status, comment), count in cancellation_comment_stats.items()
    ]
    cancellation_rating.sort(key=lambda item: (-item['count'], item['status'], item['comment']))

    # Статистика по менеджерам
    manager_stats = {}
    for order in valid_data:
        manager_id = order.get('callCenterManagerId')
        if manager_id:
            manager_stats[manager_id] = manager_stats.get(manager_id, 0) + 1

    # Динамика по дням
    daily_stats = {}
    for order in valid_data:
        created_timestamp = order.get('created')
        if created_timestamp:
            try:
                date = datetime.fromtimestamp(created_timestamp / 1000).strftime('%Y-%m-%d')
                daily_stats[date] = daily_stats.get(date, 0) + 1
            except (ValueError, TypeError, OSError):
                continue

    return {
        'total_orders': total_orders,
        'paid_orders': paid_and_processed_orders,
        'invoiced_orders': invoiced_orders,
        'conversion_rate': round(conversion_rate, 1),
        'status_stats': status_stats,
        'manager_stats': manager_stats,
        'cancellation_rating': cancellation_rating,
        'total_cancellations': total_cancellations,
        'daily_stats': daily_stats,
        'raw_data': valid_data
    }


def truncate_text(text, max_length):
    """Обрезка текста"""
    return text[:max_length] + '...' if len(text) > max_length else text


def parse_report_period():
    """Период отчёта из query: display_* и границы для БД."""
    start_param = request.args.get('start')
    end_param = request.args.get('end')

    if start_param and end_param:
        start_date = datetime.strptime(start_param, '%Y-%m-%d')
        end_date = datetime.strptime(end_param, '%Y-%m-%d')
        display_start = start_date
        display_end = end_date
        search_start = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        search_end = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    else:
        now = datetime.now()
        display_start = now.replace(day=1)
        display_end = now
        search_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        search_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

    return display_start, display_end, search_start, search_end


def build_xlsx_report_bytes(raw_orders, display_start, display_end):
    """XLSX с теми же колонками, что таблица отчёта (+ ID заявки)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "Не установлен openpyxl. Выполните: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    title = (
        f"CRM отчёт — {display_start.strftime('%d.%m.%Y')} – "
        f"{display_end.strftime('%d.%m.%Y')}"
    )
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.append([])

    headers = [
        "№",
        "ID заявки",
        "Организация",
        "ИНН",
        "Статус",
        "Комментарий для банка",
        "Менеджер",
        "Дата комментария",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, 9):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = (5, 12, 42, 14, 22, 48, 18, 18)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num = 0
    for order in raw_orders:
        if not order or not isinstance(order, dict):
            continue
        row_num += 1
        status = order.get('status', 'UNKNOWN')
        status_text = STATUS_LABELS.get(status, status)
        manager_id = order.get('callCenterManagerId', 'unknown')
        manager_name = MANAGER_NAMES.get(manager_id, f"Менеджер {manager_id}")
        oid = order.get('id', '')
        org_name = order.get('organizationName', 'Не указано')
        org_inn = order.get('organizationInn', 'Не указан')

        ws.append(
            [
                row_num,
                oid,
                org_name,
                org_inn,
                status_text,
                get_last_comment_plain(order),
                manager_name,
                get_last_comment_date_plain(order),
            ]
        )

    last_data_row = ws.max_row
    if last_data_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:H{last_data_row}"

    for r in range(header_row + 1, last_data_row + 1):
        for c in (3, 6):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def generate_html_report(analytics, start_date, end_date, db_last_update=None):
    """Генерация HTML отчета с современными стилями"""
    # Подготовка данных для графиков
    status_chart_data = []
    for status in [
        'DRAFT',
        'NEW',
        'INCOME_CREATED',
        'INCOME_PAID',
        'INCOME_PARTIALLY_PAID',
        'KKT_LINKED',
        'COMPLETED',
        'REFUND',
        'CANCELED_BY_CLIENT',
        'CANCELED_BY_BANK',
        'ARCHIVE',
    ]:
        if status in analytics['status_stats'] and analytics['status_stats'][status] > 0:
            status_chart_data.append({
                'label': STATUS_LABELS.get(status, status),
                'value': analytics['status_stats'][status]
            })

    cancellation_rating_rows = ""
    for item in analytics['cancellation_rating']:
        status = html_module.escape(str(item['status']))
        comment = html_module.escape(str(item['comment']))
        cancellation_rating_rows += f"""
                            <tr>
                                <td>{status}</td>
                                <td>{comment}</td>
                                <td>{item['count']}</td>
                                <td>{item['percent']}%</td>
                            </tr>
        """

    if not cancellation_rating_rows:
        cancellation_rating_rows = """
                            <tr>
                                <td colspan="4">За выбранный период отмен с комментариями нет.</td>
                            </tr>
        """

    # Подготовка данных для графика динамики
    daily_chart_data = []
    weekdays_ru = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    for date in sorted(analytics['daily_stats'].keys()):
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        weekday = weekdays_ru[date_obj.weekday()]
        daily_chart_data.append({
            'date': f"{date_obj.strftime('%d.%m')} ({weekday})",
            'value': analytics['daily_stats'][date]
        })

    # Генерация таблицы заявок
    table_rows = ""
    for index, order in enumerate(analytics['raw_data'], 1):
        # Проверяем что order не None и является словарем
        if not order or not isinstance(order, dict):
            continue

        status = order.get('status', 'UNKNOWN')
        status_class = STATUS_CLASSES.get(status, '')
        status_text = STATUS_LABELS.get(status, status)

        manager_id = order.get('callCenterManagerId', 'unknown')
        manager_name = MANAGER_NAMES.get(manager_id, f"Менеджер {manager_id}")

        order_id = order.get('id', 'unknown')
        org_name = order.get('organizationName', 'Не указано')
        org_inn = order.get('organizationInn', 'Не указан')

        # Обрезаем название организации
        org_name_short = truncate_text(org_name, 40)

        safe_order_id = json.dumps(order_id)
        comm_short, comm_full = get_last_comment_for_display(order)
        comment_date = get_last_comment_date_plain(order)
        table_rows += f"""
                        <tr class="clickable-row" onclick="openRequest({safe_order_id})" title="Кликните для открытия заявки">
                            <td>{index}</td>
                            <td title="{org_name}">{org_name_short}</td>
                            <td>{org_inn}</td>
                            <td><span class="status-badge {status_class}">{status_text}</span></td>
                            <td class="comment-cell" title="{comm_full}">{comm_short}</td>
                            <td>{manager_name}</td>
                            <td>{comment_date or "—"}</td>
                        </tr>
        """

    html = f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CRM Отчет - {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/3.9.1/chart.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 25px 50px rgba(0, 0, 0, 0.15);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 300;
        }}

        .period-selector {{
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            border-bottom: 1px solid #e9ecef;
        }}

        .period-form {{
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 15px;
            flex-wrap: wrap;
        }}

        .period-form input[type="date"] {{
            padding: 8px 12px;
            border: 2px solid #dee2e6;
            border-radius: 8px;
            font-size: 1rem;
        }}

        .period-form button {{
            background: #667eea;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            transition: background 0.3s;
        }}

        .period-form button:hover {{
            background: #5a6fd8;
        }}

        .period-form .btn-excel {{
            background: #217346;
        }}
        .period-form .btn-excel:hover {{
            background: #1a5c38;
        }}

        .quick-periods {{
            margin-top: 10px;
            display: flex;
            justify-content: center;
            gap: 10px;
            flex-wrap: wrap;
        }}

        .quick-periods button {{
            background: #6c757d;
            color: white;
            padding: 5px 12px;
            border: none;
            border-radius: 15px;
            cursor: pointer;
            font-size: 0.9em;
            transition: background 0.3s;
        }}

        .quick-periods button:hover {{
            background: #5a6268;
        }}

        .period {{
            background: #f8f9fa;
            padding: 15px;
            text-align: center;
            font-size: 1.2em;
            color: #495057;
            border-bottom: 1px solid #e9ecef;
        }}

        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
        }}

        .stat-card {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
            text-align: center;
            transition: transform 0.3s;
        }}

        .stat-card:hover {{
            transform: translateY(-5px);
        }}

        .stat-value {{
            font-size: 2.5em;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 5px;
        }}

        .stat-label {{
            color: #6c757d;
            font-size: 0.9em;
            font-weight: 500;
        }}

        .chart-section {{
            padding: 30px;
            background: #f8f9fa;
        }}

        .chart-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin-bottom: 30px;
        }}

        .chart-container {{
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
        }}

        .chart-title {{
            font-size: 1.3em;
            font-weight: 600;
            color: #2c3e50;
            margin-bottom: 20px;
            text-align: center;
        }}

        .table-section {{
            padding: 30px;
        }}

        .table-container {{
            background: white;
            border-radius: 15px;
            overflow: hidden;
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
        }}

        .table-header {{
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: white;
            padding: 20px;
            font-size: 1.2em;
            font-weight: 600;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
        }}

        th, td {{
            padding: 12px 15px;
            text-align: left;
            border-bottom: 1px solid #e9ecef;
        }}

        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #495057;
            font-size: 0.9em;
        }}

        tr:hover {{
            background: #f8f9fa;
            cursor: pointer;
        }}

        .clickable-row {{
            cursor: pointer;
            transition: background-color 0.2s;
        }}

        .clickable-row:hover {{
            background: #e9ecef !important;
        }}

        .status-badge {{
            padding: 4px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: 600;
        }}

        .status-new {{ background: #e3f2fd; color: #1976d2; }}
        .status-draft {{ background: #f3e5f5; color: #7b1fa2; }}
        .status-income-created {{ background: #fff3e0; color: #f57c00; }}
        .status-income-paid {{ background: #e8f5e8; color: #388e3c; }}
        .status-income-partially-paid {{ background: #fff8e1; color: #f9a825; }}
        .status-kkt-linked {{ background: #e8f5e8; color: #2e7d32; }}
        .status-completed {{ background: #e8f5e8; color: #2e7d32; }}
        .status-refund {{ background: #e3f2fd; color: #1565c0; }}
        .status-canceled {{ background: #ffebee; color: #d32f2f; }}
        .status-canceled-by-bank {{ background: #ffebee; color: #b71c1c; }}
        .status-archive {{ background: #eceff1; color: #455a64; }}

        .comment-cell {{
            font-size: 0.9em;
            max-width: 320px;
            line-height: 1.35;
            color: #374151;
        }}

        @media print {{
            .period-selector {{ display: none; }}
            body {{ background: white; }}
            .container {{ box-shadow: none; }}
        }}

        @media (max-width: 768px) {{
            .chart-grid {{ grid-template-columns: 1fr; }}
            .period-form {{ flex-direction: column; }}
        }}
    </style>
</head>
<body>

    <div class="container">
        <div class="header">
            <h1>📊 CRM Отчет</h1>
            <p>Управленческая отчетность и аналитика</p>
        </div>

        <div class="period-selector">
            <form class="period-form" method="GET">
                <label>с:</label>
                <input type="date" name="start" value="{start_date.strftime('%Y-%m-%d')}" required>
                <label>по:</label>
                <input type="date" name="end" value="{end_date.strftime('%Y-%m-%d')}" required>
                <button type="submit">📊 Сформировать отчет</button>
                <button type="button" class="btn-excel" onclick="downloadReportExcel()">📥 Скачать Excel</button>
            </form>
            <div class="quick-periods">
                <button onclick="setQuickPeriod('today')">Сегодня</button>
                <button onclick="setQuickPeriod('yesterday')">Вчера</button>
                <button onclick="setQuickPeriod('week')">Неделя</button>
                <button onclick="setQuickPeriod('month')">Месяц</button>
                <button onclick="setQuickPeriod('quarter')">Квартал</button>
            </div>
        </div>

        <div class="period">
            <strong>Период:</strong> {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}
        </div>
        <div class="period" style="font-size:0.95em;color:#6c757d;">
            В отчёт входят только заявки, <strong>созданные</strong> в выбранном периоде (по дате появления в CRM).
            Последняя синхронизация данных по ним в БД: <strong>{db_last_update.strftime('%d.%m.%Y %H:%M') if db_last_update else '—'}</strong>.
            <a href="/sync?next=/report" style="margin-left:8px;color:#667eea;">Обновить из CRM сейчас</a>
        </div>

        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-value">{analytics['total_orders']}</div>
                <div class="stat-label">Всего заявок</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{analytics['paid_orders']}</div>
                <div class="stat-label">Оплаченных</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{analytics['conversion_rate']}%</div>
                <div class="stat-label">Конверсия</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{analytics['invoiced_orders']}</div>
                <div class="stat-label">Ожидаем оплат</div>
            </div>
        </div>

        <div class="chart-section">
            <div class="chart-grid">
                <div class="chart-container">
                    <div class="chart-title">Распределение по статусам</div>
                    <canvas id="statusChart"></canvas>
                </div>
                <div class="chart-container">
                    <div class="chart-title">Рейтинг причин отмен</div>
                    <div style="overflow-x: auto;">
                        <table>
                            <thead>
                                <tr>
                                    <th>Статус</th>
                                    <th>Комментарий для банка</th>
                                    <th>Количество</th>
                                    <th>Доля отмен</th>
                                </tr>
                            </thead>
                            <tbody>
                                {cancellation_rating_rows}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            <div class="chart-container">
                <div class="chart-title">Динамика создания заявок</div>
                <canvas id="timelineChart"></canvas>
            </div>
        </div>

        <div class="table-section">
            <div class="table-container">
                <div class="table-header">Детальная информация по заявкам</div>
                <table>
                    <thead>
                        <tr>
                            <th>№</th>
                            <th>Организация</th>
                            <th>ИНН</th>
                            <th>Статус</th>
                            <th>Комментарий для банка</th>
                            <th>Менеджер</th>
                            <th>Дата комментария</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
        // Функция для открытия заявки в новом окне
        function openRequest(orderId) {{
            if (orderId && orderId !== 'unknown') {{
                const url = `{CRM_CONFIG['admin_url']}?id=${{orderId}}`;
                window.open(url, '_blank');
            }} else {{
                alert('ID заявки не найден');
            }}
        }}

        function downloadReportExcel() {{
            const form = document.querySelector('.period-form');
            const start = form.querySelector('input[name="start"]').value;
            const end = form.querySelector('input[name="end"]').value;
            if (!start || !end) {{
                alert('Укажите период (даты с / по)');
                return;
            }}
            const q = new URLSearchParams({{ start, end }}).toString();
            window.location.href = '/report/export?' + q;
        }}

        // Дата в локальном календаре (не UTC — иначе «Сегодня» съезжает на соседний день)
        function toLocalYMD(d) {{
            const y = d.getFullYear();
            const m = String(d.getMonth() + 1).padStart(2, '0');
            const day = String(d.getDate()).padStart(2, '0');
            return y + '-' + m + '-' + day;
        }}

        // Функция для быстрого выбора периода
        function setQuickPeriod(period) {{
            const now = new Date();
            let startDate, endDate;

            switch(period) {{
                case 'today':
                    startDate = endDate = now;
                    break;
                case 'yesterday': {{
                    const yesterday = new Date(now);
                    yesterday.setDate(yesterday.getDate() - 1);
                    startDate = endDate = yesterday;
                    break;
                }}
                case 'week':
                    endDate = now;
                    startDate = new Date(now);
                    startDate.setDate(startDate.getDate() - 7);
                    break;
                case 'month':
                    endDate = now;
                    startDate = new Date(now);
                    startDate.setMonth(startDate.getMonth() - 1);
                    break;
                case 'quarter':
                    endDate = now;
                    startDate = new Date(now);
                    startDate.setMonth(startDate.getMonth() - 3);
                    break;
            }}

            document.querySelector('input[name="start"]').value = toLocalYMD(startDate);
            document.querySelector('input[name="end"]').value = toLocalYMD(endDate);
        }}

        // Данные для графиков
        const statusData = {json.dumps(status_chart_data)};
        const dailyData = {json.dumps(daily_chart_data)};

        const chartColors = [
            '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40'
        ];

        // График статусов (Chart.js падает на полностью пустых данных)
        if (statusData.length > 0) {{
            const statusCtx = document.getElementById('statusChart').getContext('2d');
            new Chart(statusCtx, {{
                type: 'doughnut',
                data: {{
                    labels: statusData.map(item => item.label),
                    datasets: [{{
                        data: statusData.map(item => item.value),
                        backgroundColor: chartColors
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{
                        legend: {{ position: 'bottom' }}
                    }}
                }}
            }});
        }}

        if (dailyData.length > 0) {{
            const timelineCtx = document.getElementById('timelineChart').getContext('2d');
            new Chart(timelineCtx, {{
                type: 'line',
                data: {{
                    labels: dailyData.map(item => item.date),
                    datasets: [{{
                        label: 'Количество заявок',
                        data: dailyData.map(item => item.value),
                        borderColor: '#667eea',
                        backgroundColor: 'rgba(102, 126, 234, 0.1)',
                        fill: true,
                        tension: 0.4
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{ legend: {{ display: false }} }},
                    scales: {{ y: {{ beginAtZero: true }} }}
                }}
            }});
        }}
    </script>
</body>
</html>
    """
    return html


@app.route('/report/export')
def export_report_xlsx():
    """Выгрузка таблицы отчёта в Excel (те же start/end, что у /report)."""
    try:
        display_start, display_end, search_start, search_end = parse_report_period()
        data = get_requests_from_db(search_start, search_end)
        analytics = analyze_data(data)
        blob = build_xlsx_report_bytes(analytics['raw_data'], display_start, display_end)
        fname = (
            f"crm_report_{display_start.strftime('%Y-%m-%d')}_"
            f"{display_end.strftime('%Y-%m-%d')}.xlsx"
        )
        return Response(
            blob,
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f'attachment; filename="{fname}"',
            },
        )
    except RuntimeError as e:
        return Response(str(e), mimetype="text/plain; charset=utf-8", status=500)
    except Exception as e:
        logger.exception("Ошибка выгрузки Excel")
        return Response(
            f"Ошибка: {e}",
            mimetype="text/plain; charset=utf-8",
            status=500,
        )


@app.route('/report')
def generate_report():
    """API для генерации отчета"""
    display_start, display_end, search_start, search_end = parse_report_period()
    start_date = display_start
    end_date = display_end

    try:
        # Получаем данные из БД используя временной интервал
        data = get_requests_from_db(search_start, search_end)
        db_last = get_last_db_update_in_period(search_start, search_end)

        # Анализируем
        analytics = analyze_data(data)

        # Генерируем HTML (передаем даты для отображения)
        html = generate_html_report(analytics, display_start, display_end, db_last)

        return Response(html, mimetype='text/html')

    except Exception as e:
        logger.exception("Ошибка генерации отчета")
        err_html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ошибка отчёта</title>
    <style>
        body {{ font-family: system-ui, sans-serif; max-width: 640px; margin: 40px auto; padding: 0 16px; }}
        h1 {{ color: #c0392b; }}
        pre {{ background: #f4f4f4; padding: 12px; overflow: auto; border-radius: 8px; }}
        a {{ color: #667eea; }}
    </style>
</head>
<body>
    <h1>Ошибка генерации отчёта</h1>
    <p>Проверьте, что PostgreSQL запущен и база <code>crm_reports</code> доступна с учётными данными из <code>config.py</code>.</p>
    <pre>{html_module.escape(str(e))}</pre>
    <p><a href="/report">Повторить</a></p>
</body>
</html>"""
        return Response(err_html, mimetype="text/html", status=500)


if __name__ == "__main__":
    from config import WEB_CONFIG

    app.run(**WEB_CONFIG)